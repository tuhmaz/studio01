import openpyxl
from pathlib import Path

FILE_NAME = "Tourplan 2025 Kopie.xlsx"
OUTPUT_FILE = "excel_output_full.txt"
MAX_ROWS_TO_PRINT = None  # ضع رقمًا مثل 100 إذا أردت حدًا معينًا


def get_fill_color(cell):
    """
    استخراج لون تعبئة الخلية إن وجد.
    """
    fill = cell.fill

    if fill is None or fill.fill_type is None:
        return None

    color = fill.fgColor

    if color is None:
        return None

    if color.type == "rgb":
        return color.rgb
    elif color.type == "indexed":
        return f"indexed:{color.indexed}"
    elif color.type == "theme":
        tint = getattr(color, "tint", None)
        return f"theme:{color.theme}, tint:{tint}"
    elif color.type == "auto":
        return "auto"

    return None


def build_merged_map(ws):
    """
    بناء خريطة للخلايا المدموجة بحيث كل خلية ضمن الدمج
    تشير إلى الخلية الأساسية وقيمتها.
    """
    merged_map = {}

    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        top_left_cell = ws.cell(min_row, min_col)
        top_left_value = top_left_cell.value
        range_str = str(merged_range)

        for row in ws.iter_rows(
            min_row=merged_range.min_row,
            max_row=merged_range.max_row,
            min_col=merged_range.min_col,
            max_col=merged_range.max_col,
        ):
            for cell in row:
                merged_map[cell.coordinate] = {
                    "range": range_str,
                    "top_left": top_left_cell.coordinate,
                    "value": top_left_value,
                }

    return merged_map


def serialize_cell(cell, merged_map):
    """
    تجهيز معلومات الخلية.
    """
    raw_value = cell.value
    display_value = raw_value
    fill_color = get_fill_color(cell)
    is_merged = False
    merged_range = None
    merged_source = None

    if cell.coordinate in merged_map:
        merged_info = merged_map[cell.coordinate]
        is_merged = True
        merged_range = merged_info["range"]
        merged_source = merged_info["top_left"]

        if raw_value is None:
            display_value = merged_info["value"]

    return {
        "cell": cell.coordinate,
        "raw_value": raw_value,
        "display_value": display_value,
        "fill_color": fill_color,
        "is_merged": is_merged,
        "merged_range": merged_range,
        "merged_source": merged_source,
    }


def row_has_visible_data(serialized_row):
    """
    نعتبر الصف مهمًا إذا احتوى قيمة أو لون أو دمج.
    """
    for item in serialized_row:
        if (
            item["display_value"] is not None
            or item["fill_color"] is not None
            or item["is_merged"]
        ):
            return True
    return False


def main():
    file_path = Path(FILE_NAME)

    if not file_path.exists():
        print(f"ERROR: File not found -> {FILE_NAME}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    with open(OUTPUT_FILE, "w", encoding="utf-8") as out:
        out.write("=== SHEETS ===\n")
        out.write(f"{wb.sheetnames}\n\n")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            merged_map = build_merged_map(ws)

            out.write(f"=== SHEET: {sheet_name} ===\n")
            out.write(f"Dimensions: {ws.dimensions}\n")
            out.write(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}\n")
            out.write("\n")

            printed_rows = 0

            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                if MAX_ROWS_TO_PRINT is not None and printed_rows >= MAX_ROWS_TO_PRINT:
                    break

                serialized_row = [serialize_cell(cell, merged_map) for cell in row]

                if not row_has_visible_data(serialized_row):
                    continue

                visible_cells = [
                    item for item in serialized_row
                    if (
                        item["display_value"] is not None
                        or item["fill_color"] is not None
                        or item["is_merged"]
                    )
                ]

                out.write(f"Row {row_idx}:\n")

                for item in visible_cells:
                    out.write(
                        f"  {item['cell']} | "
                        f"value={item['display_value']} | "
                        f"raw={item['raw_value']} | "
                        f"fill={item['fill_color']} | "
                        f"merged={item['is_merged']} | "
                        f"range={item['merged_range']} | "
                        f"source={item['merged_source']}\n"
                    )

                out.write("\n")
                printed_rows += 1

    print(f"Done. Results saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()